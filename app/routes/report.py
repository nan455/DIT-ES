from flask import Blueprint, request, jsonify, send_file, current_app
from app.utils.database import with_db_connection
from app.utils.report_config import REPORT_CONFIG
from flask import session
from app.utils.validators import resolve_year
from collections import OrderedDict
import io, os, datetime

# ======================================================
# Blueprint
# ======================================================
report_bp = Blueprint("report", __name__, url_prefix="/api/report")

# ======================================================
# Helpers
# ======================================================
ROMAN = ["I","II","III","IV","V","VI","VII","VIII","IX","X"]

def roman(n):
    return ROMAN[n-1] if n <= len(ROMAN) else str(n)

def safe(v):
    return 0 if v is None else v

# THOUSAND SEPARATOR
def fmt(n):
    return "{:,}".format(int(n)) if n is not None else "0"


# ======================================================
# Load Regions
# ======================================================
def load_regions(cursor):
    cursor.execute("""
        SELECT region_id, region_desc
        FROM tbl_region_master
        ORDER BY region_id
    """)
    return cursor.fetchall()


# ======================================================
# Year Label
# ======================================================
def get_year_label(cursor, year_id):
    cursor.execute(
        "SELECT year_desc FROM tbl_year_master WHERE year_id=%s",
        (year_id,)
    )
    r = cursor.fetchone()
    return r["year_desc"] if r else str(year_id)


# ======================================================
# SQL Builder
# ======================================================
def build_report_query(cursor, table, year):

    if table not in REPORT_CONFIG:
        raise Exception("Unsupported report table")

    cfg = REPORT_CONFIG[table]
    regions = load_regions(cursor)
    has_item = cfg["item"] is not None

    select_item = (
        f", i.{cfg['item']['label']} AS item" if has_item else ", NULL AS item"
    )

    join_item = (
        f"JOIN {cfg['item']['table']} i ON {cfg['item']['join']}"
        if has_item else ""
    )

    region_sql = []
    for r in regions:
        region_sql.append(
            f"""SUM(CASE WHEN c.region_id={r['region_id']}
                THEN c.{cfg['value_column']} ELSE 0 END)
                AS `{r['region_desc']}`"""
        )

    region_sql.append(f"SUM(c.{cfg['value_column']}) AS Total")

    sql = f"""
        SELECT g.{cfg['group']['label']} AS grp
        {select_item},
        {",".join(region_sql)}
        FROM {table} c
        JOIN {cfg['group']['table']} g
          ON {cfg['group']['join']}
        {join_item}
        WHERE c.{cfg['year_column']}=%s
        GROUP BY grp {", item" if has_item else ""}
        ORDER BY grp {", item" if has_item else ""}
    """

    return sql, [year], regions


# ======================================================
# Build JSON
# ======================================================
def build_report(rows, regions):

    region_names = [r["region_desc"] for r in regions] + ["Total"]

    report = OrderedDict()
    grand = {r:0 for r in region_names}

    idx = 1
    for r in rows:
        grp = r["grp"]

        if grp not in report:
            report[grp] = {
                "roman": roman(idx),
                "rows": [],
                "total": {k:0 for k in region_names}
            }
            idx += 1

        row = {"item": r.get("item") or ""}

        for k in region_names:
            val = safe(r[k])
            row[k] = val
            report[grp]["total"][k] += val
            grand[k] += val

        report[grp]["rows"].append(row)

    return {
        "groups": report,
        "grand_total": grand,
        "regions": region_names[:-1]
    }


# ======================================================
# JSON API
# ======================================================
@report_bp.route("/<table>")
@with_db_connection
def get_report(cursor, conn, table):

    year = request.args.get("year")
    if not year:
        return jsonify({"error":"year required"}),400

    year = resolve_year(cursor, year)

    sql, params, regions = build_report_query(cursor, table, year)
    cursor.execute(sql, params)
    rows = cursor.fetchall()

    return jsonify(build_report(rows, regions))


# ======================================================
# REPORT LABEL
# ======================================================
@report_bp.route("/table_label/<table>")
def get_table_label(table):
    cfg = REPORT_CONFIG.get(table)
    return jsonify({"label": cfg.get("title", table)})


# ======================================================
# PDF DOWNLOAD
# ======================================================
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

@report_bp.route("/<table>/pdf")
@with_db_connection
def download_pdf(cursor, conn, table):

    year_param = request.args.get("year")
    year_id = resolve_year(cursor, year_param)
    year_label = get_year_label(cursor, year_id)

    report_title = REPORT_CONFIG.get(table, {}).get("title", table)

    sql, params, regions = build_report_query(cursor, table, year_id)
    cursor.execute(sql, params)
    rows = cursor.fetchall()
    data = build_report(rows, regions)

    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=30,
        rightMargin=30,
        topMargin=40,
        bottomMargin=40
    )

    styles = getSampleStyleSheet()

    H1 = ParagraphStyle("H1", parent=styles["Normal"], alignment=1, fontSize=14, fontName="Helvetica-Bold")
    H2 = ParagraphStyle("H2", parent=styles["Normal"], alignment=1, fontSize=11)
    H3 = ParagraphStyle("H3", parent=styles["Normal"], alignment=1, fontSize=10)

    elements = []

    elements.append(Paragraph("DIRECTORATE OF ECONOMICS AND STATISTICS", H1))
    dept = session.get("department", "")
    elements.append(Paragraph(dept, H2))
    elements.append(Paragraph(report_title, H2))
    elements.append(Paragraph(f"Year : {year_label}", H3))
    elements.append(Spacer(1, 12))

    headers = ["Sl.No","Item"] + data["regions"] + ["Total"]
    table_data = [headers]

    for grp, g in data["groups"].items():

        table_data.append([f"{g['roman']}. {grp}"] + [""]*(len(headers)-1))

        for r in g["rows"]:
            table_data.append(
                ["", r["item"]] +
                [fmt(r[n]) for n in data["regions"]] +
                [fmt(r["Total"])]
            )

        table_data.append(
            ["", f"Total {grp}"] +
            [fmt(g["total"][n]) for n in data["regions"]] +
            [fmt(g["total"]["Total"])]
        )

    table_data.append(
        ["","GRAND TOTAL"] +
        [fmt(data["grand_total"][n]) for n in data["regions"]] +
        [fmt(data["grand_total"]["Total"])]
    )

    tbl = Table(table_data, repeatRows=1)

    tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1F4E79")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("ALIGN",(0,0),(-1,0),"CENTER"),

        ("GRID",(0,0),(-1,-1),0.5,colors.grey),
        ("ALIGN",(2,1),(-1,-1),"RIGHT"),

        ("BACKGROUND",(0,-1),(-1,-1),colors.lightgrey),
        ("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold")
    ]))

    elements.append(tbl)

    def footer(canvas, doc):
        canvas.drawRightString(820,20,f"Generated on {datetime.datetime.now().strftime('%d/%m/%Y')}")
        canvas.drawCentredString(420,20,f"Page {doc.page}")

    doc.build(elements,onFirstPage=footer,onLaterPages=footer)

    buffer.seek(0)
    return send_file(buffer, download_name="report.pdf", as_attachment=True)


# ======================================================
# EXCEL DOWNLOAD
# ======================================================
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

@report_bp.route("/<table>/excel")
@with_db_connection
def download_excel(cursor, conn, table):

    year_param = request.args.get("year")
    year_id = resolve_year(cursor, year_param)
    year_label = get_year_label(cursor, year_id)

    report_title = REPORT_CONFIG.get(table, {}).get("title", table)

    sql, params, regions = build_report_query(cursor, table, year_id)
    cursor.execute(sql, params)
    rows = cursor.fetchall()
    data = build_report(rows, regions)

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    headers = ["Sl.No","Item"] + data["regions"] + ["Total"]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    white_bold = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)
    right = Alignment(horizontal="right")
    center = Alignment(horizontal="center")

    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    ws.append(["DIRECTORATE OF ECONOMICS AND STATISTICS"])
    ws.append(["Animal Husbandry and Animal Welfare"])
    ws.append([report_title])
    ws.append([f"Year : {year_label}"])

    for r in range(1,5):
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=len(headers))
        ws.cell(r,1).font=bold
        ws.cell(r,1).alignment=center

    ws.append(headers)

    for c in ws[5]:
        c.font=white_bold
        c.fill=header_fill
        c.alignment=center
        c.border=thin

    for grp, g in data["groups"].items():

        ws.append([f"{g['roman']}. {grp}"])
        ws.merge_cells(start_row=ws.max_row,start_column=1,
                       end_row=ws.max_row,end_column=len(headers))
        ws.cell(ws.max_row,1).font=bold

        for r in g["rows"]:
            ws.append(
                ["", r["item"]] +
                [fmt(r[n]) for n in data["regions"]] +
                [fmt(r["Total"])]
            )

        ws.append(
            ["", f"Total {grp}"] +
            [fmt(g["total"][n]) for n in data["regions"]] +
            [fmt(g["total"]["Total"])]
        )
        ws.cell(ws.max_row,2).font=bold

    ws.append(
        ["","GRAND TOTAL"] +
        [fmt(data["grand_total"][n]) for n in data["regions"]] +
        [fmt(data["grand_total"]["Total"])]
    )
    ws.cell(ws.max_row,2).font=bold

    for row in ws.iter_rows():
        for c in row:
            c.border=thin
            if c.column>2:
                c.alignment=right

    ws.freeze_panes="A6"

    for i in range(1,len(headers)+1):
        maxlen=0
        for cell in ws[get_column_letter(i)]:
            if cell.value:
                maxlen=max(maxlen,len(str(cell.value)))
        ws.column_dimensions[get_column_letter(i)].width=maxlen+3

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    return send_file(out, download_name="report.xlsx", as_attachment=True)
